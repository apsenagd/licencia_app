from flask import Flask, render_template, request, redirect, jsonify, send_file
import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import os

app = Flask(__name__)

# CONEXIÓN MYSQL

def obtener_conexion():


    return mysql.connector.connect(
    host=os.getenv("DB_HOST"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD"),
    database=os.getenv("DB_NAME"),
    port=int(os.getenv("DB_PORT", 3306))

)

# CERRAR CONEXIONES
def cerrar_conexion(cursor, conexion):
    try:
        if cursor:
            cursor.close()
    except:
        pass

    try:
        if conexion and conexion.is_connected():
            conexion.close()
    except:
        pass

# ---------------------------
# FORMULARIO 1
# ---------------------------
@app.route('/')
def inicio():
    return render_template('encargado.html')


# ---------------------------
# FORMULARIO 2
# ---------------------------
@app.route('/formulario_licencia', methods=['POST'])
def formulario_licencia():

    encargado = request.form['encargado']
    area = request.form['area']
    fecha_hora = request.form.get('fecha_hora')
    solicitante = request.form['solicitante']

    fecha_solicitud = None
    if fecha_hora:
        try:
            fecha_solicitud = datetime.fromisoformat(fecha_hora).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            fecha_solicitud = fecha_hora

    conexion = obtener_conexion()
    cursor = conexion.cursor()
    sql = '''
    INSERT INTO solicitudes (jefe_area, area_departamento, solicitante, fecha_solicitud)
    VALUES (%s, %s, %s, %s)
    '''
    cursor.execute(sql, (encargado, area, solicitante, fecha_solicitud))
    conexion.commit()
    id_solicitud = cursor.lastrowid

    cerrar_conexion(cursor, conexion)

    return render_template(
        'formulario_licencia.html',
        encargado=encargado,
        area=area,
        fecha_hora=fecha_hora,
        solicitante=solicitante,
        id_solicitud=id_solicitud,
        subarea=''
    )


# ---------------------------
# VISTA 2
# ---------------------------
@app.route('/vista2', methods=['POST'])
def vista2():
    # Preserve multi-select fields (tipo_info_sensible) when posting between steps
    datos = request.form.to_dict()
    tipo_vals = request.form.getlist('tipo_info_sensible')
    if tipo_vals:
        datos['tipo_info_sensible'] = ', '.join(tipo_vals)
    else:
        # fallback: single value or empty
        datos['tipo_info_sensible'] = request.form.get('tipo_info_sensible', '')
    return render_template('vista2.html', **datos)


# ---------------------------
# VISTA 3
# ---------------------------
@app.route('/vista3', methods=['POST'])
def vista3():
    # Preserve multi-select fields (tipo_info_sensible) when posting between steps
    datos = request.form.to_dict()
    tipo_vals = request.form.getlist('tipo_info_sensible')
    if tipo_vals:
        datos['tipo_info_sensible'] = ', '.join(tipo_vals)
    else:
        datos['tipo_info_sensible'] = request.form.get('tipo_info_sensible', '')
    return render_template('vista3.html', **datos)


# ---------------------------
# GUARDAR INFORMACIÓN
# ---------------------------
@app.route('/guardar', methods=['POST'])
def guardar():

    # DATOS FORMULARIO
    encargado = request.form['encargado']
    area = request.form['area']
    solicitante = request.form['solicitante']

    licencia = request.form['licencia']
    tipo_herramienta = request.form.get('tipo_herramienta')
    licenciamiento = request.form.get('licenciamiento')
    proveedor = request.form.get('proveedor')
    frecuencia = request.form.get('frecuencia')
    vigencia = request.form['vigencia']
    descripcion = request.form.get('uso')

    usuarios_area = request.form.get('usuarios_area')
    roles = request.form.get('roles')
    gestiona_roles = request.form.get('gestiona_roles')
    gestiona_usuarios = request.form.get('gestiona_usuarios')
    info_sensible = request.form.get('info_sensible')
    tipo_info_sensible_values = request.form.getlist('tipo_info_sensible')
    # Si el valor viene como una sola cadena separada por comas (desde pasos intermedios), dividirla
    if len(tipo_info_sensible_values) == 1 and tipo_info_sensible_values[0] and ',' in tipo_info_sensible_values[0]:
        tipo_info_sensible_values = [v.strip() for v in tipo_info_sensible_values[0].split(',')]
    info_sensible_otro = request.form.get('info_sensible_otro')
    if 'Otro' in tipo_info_sensible_values and info_sensible_otro:
        tipo_info_sensible_values = [value for value in tipo_info_sensible_values if value != 'Otro'] + [info_sensible_otro]
    tipo_info_sensible = ', '.join(tipo_info_sensible_values) if tipo_info_sensible_values else None
    compliance = request.form.get('compliance_validado')

    copias = request.form.get('copias_seguridad')
    tipo_copias = request.form.get('tipo_copias')
    frecuencia_copias = request.form.get('frecuencia_copias')
    ubicacion = request.form.get('ubicacion_copias')
    integraciones = request.form.get('integracion_sistemas')
    detalle_integracion = request.form.get('sistemas_integracion')
    impacto = request.form.get('criticidad')
    procesos = request.form.get('procesos_apoyo')
    contacto = request.form.get('contacto_proveedor')

    conexion = obtener_conexion()
    cursor = conexion.cursor()
    id_solicitud = request.form.get('id_solicitud')
    if id_solicitud:
        id_solicitud = int(id_solicitud)
    else:
        fecha_solicitud = request.form.get('fecha_hora')
        if fecha_solicitud:
            try:
                fecha_solicitud = datetime.fromisoformat(fecha_solicitud).strftime('%Y-%m-%d %H:%M:%S')
            except ValueError:
                fecha_solicitud = None

        sql1 = '''
        INSERT INTO solicitudes (jefe_area, area_departamento, solicitante, fecha_solicitud)
        VALUES (%s, %s, %s, %s)
        '''
        cursor.execute(sql1, (encargado, area, solicitante, fecha_solicitud))
        id_solicitud = cursor.lastrowid

    # -------- 2. licencias --------
    sql2 = '''
    INSERT INTO licencias (
        id_solicitud,
        nombre_licencia,
        descripcion,
        tipo_herramienta,
        licenciamiento,
        proveedor,
        frecuencia_uso,
        fecha_vencimiento
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    '''
    cursor.execute(sql2, (
        id_solicitud,
        licencia,
        descripcion,
        tipo_herramienta,
        licenciamiento,
        proveedor,
        frecuencia,
        vigencia
    ))
    id_licencia = cursor.lastrowid

    # -------- 3. datos adicionales --------
    sql3 = '''
    INSERT INTO datos_adicionales (
        id_licencia,
        cantidad_usuarios,
        maneja_roles,
        gestor_roles,
        gestor_usuarios,
        maneja_info_sensible,
        tipo_info_sensible,
        validacion_compliance
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    '''
    cursor.execute(sql3, (
        id_licencia,
        usuarios_area,
        roles,
        gestiona_roles,
        gestiona_usuarios,
        info_sensible,
        tipo_info_sensible,
        compliance
    ))

    # -------- 4. copias seguridad --------
    sql4 = '''
    INSERT INTO copias_seguridad (
        id_licencia,
        realiza_copias,
        tipo_copia,
        frecuencia_copia,
        ubicacion_copia,
        tiene_integraciones,
        detalle_integraciones,
        impacto_falla,
        procesos_relacionados,
        contacto_proveedor
    )
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    '''
    cursor.execute(sql4, (
        id_licencia,
        copias,
        tipo_copias,
        frecuencia_copias,
        ubicacion,
        integraciones,
        detalle_integracion,
        impacto,
        procesos,
        contacto
    ))

    conexion.commit()
    cerrar_conexion(cursor, conexion)

    return redirect(f'/solicitud/{id_solicitud}/resumen')


# ---------------------------
# VISTA
# ---------------------------
@app.route('/vista')
def vista():

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    cursor.execute('''
        SELECT 
            l.id_licencia,

            s.area_departamento AS area,
            s.jefe_area AS encargado,
            s.solicitante,

            l.nombre_licencia AS licencia,
            l.descripcion AS uso,
            l.fecha_vencimiento AS vigencia

        FROM solicitudes s
        JOIN licencias l ON s.id_solicitud = l.id_solicitud
    ''')

    datos = cursor.fetchall()

    cerrar_conexion(cursor, conexion)

    return render_template('vista_licencias.html', licencias=datos)


@app.route('/solicitud/<int:id_solicitud>/licencia/nueva')
def nueva_licencia(id_solicitud):
    
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT jefe_area, area_departamento, solicitante, fecha_solicitud
        FROM solicitudes
        WHERE id_solicitud = %s
    ''', (id_solicitud,))
    solicitud = cursor.fetchone()

    cerrar_conexion(cursor, conexion)
    if not solicitud:
        return redirect('/')

    return render_template('formulario_licencia.html',
        encargado=solicitud['jefe_area'],
        area=solicitud['area_departamento'],
        fecha_hora=solicitud['fecha_solicitud'],
        solicitante=solicitud['solicitante'],
        id_solicitud=id_solicitud,
        subarea=''
    )


@app.route('/solicitud/<int:id_solicitud>/resumen')
def solicitud_resumen(id_solicitud):
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT id_solicitud, jefe_area, area_departamento, solicitante, fecha_solicitud
        FROM solicitudes
        WHERE id_solicitud = %s
    ''', (id_solicitud,))
    solicitud = cursor.fetchone()
    if not solicitud:
        return redirect('/')

    cursor.execute('''
        SELECT id_licencia, nombre_licencia, descripcion, fecha_vencimiento
        FROM licencias
        WHERE id_solicitud = %s
    ''', (id_solicitud,))
    licencias = cursor.fetchall()

    cerrar_conexion(cursor, conexion)

    return render_template('solicitud_resumen.html', solicitud=solicitud, licencias=licencias)


@app.route('/solicitud/<int:id_solicitud>/finalizado')
def finalizado_solicitud(id_solicitud):
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT id_solicitud, jefe_area, area_departamento, solicitante, fecha_solicitud
        FROM solicitudes
        WHERE id_solicitud = %s
    ''', (id_solicitud,))
    solicitud = cursor.fetchone()
    if not solicitud:
        return redirect('/')
    
    cursor.execute('''
        SELECT COUNT(*) as total_licencias
        FROM licencias
        WHERE id_solicitud = %s
    ''', (id_solicitud,))
    result = cursor.fetchone()
    total_licencias = result['total_licencias'] if result else 0

    cerrar_conexion(cursor, conexion)
    
    return render_template('final.html', solicitud=solicitud, total_licencias=total_licencias)


@app.route('/solicitudes')
def solicitudes():
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT
            s.id_solicitud,
            s.area_departamento AS area,
            s.jefe_area AS encargado,
            s.solicitante,
            s.fecha_solicitud,
            COUNT(l.id_licencia) AS total_licencias
        FROM solicitudes s
        LEFT JOIN licencias l ON s.id_solicitud = l.id_solicitud
        GROUP BY s.id_solicitud
        ORDER BY s.fecha_solicitud DESC
    ''')
    datos = cursor.fetchall()

    cerrar_conexion(cursor, conexion)

    return render_template('solicitudes.html', solicitudes=datos)


# ---------------------------
# LUPA

@app.route('/vista_global')
def vista_global():

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT
            s.area_departamento AS area,
            s.jefe_area AS encargado,
            s.solicitante,
            s.fecha_solicitud AS fecha_solicitud,
            l.id_licencia,
            l.nombre_licencia AS nombre_licencia,
            l.descripcion AS descripcion,
            l.fecha_vencimiento AS fecha_vencimiento
        FROM licencias l
        JOIN solicitudes s ON l.id_solicitud = s.id_solicitud
        ORDER BY s.area_departamento, l.nombre_licencia
    ''')
    rows = cursor.fetchall()

    cerrar_conexion(cursor, conexion)

    grouped = {}
    for r in rows:
        area = r.get('area') or 'Sin área'
        grouped.setdefault(area, []).append(r)

    return render_template('vista_global.html', grouped=grouped)

@app.route('/exportar_licencias')
def exportar_licencias():
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('''
        SELECT
            l.id_licencia,
            l.nombre_licencia,
            l.descripcion,
            l.tipo_herramienta,
            l.licenciamiento,
            l.proveedor,
            l.frecuencia_uso,
            l.fecha_vencimiento,
            s.id_solicitud,
            s.area_departamento,
            s.jefe_area,
            s.solicitante,
            s.fecha_solicitud,
            da.cantidad_usuarios,
            da.maneja_roles,
            da.gestor_roles,
            da.gestor_usuarios,
            da.maneja_info_sensible,
            da.tipo_info_sensible,
            da.validacion_compliance,
            cs.realiza_copias,
            cs.tipo_copia,
            cs.frecuencia_copia,
            cs.ubicacion_copia,
            cs.tiene_integraciones,
            cs.detalle_integraciones,
            cs.impacto_falla,
            cs.procesos_relacionados,
            cs.contacto_proveedor
        FROM licencias l
        LEFT JOIN solicitudes s ON l.id_solicitud = s.id_solicitud
        LEFT JOIN datos_adicionales da ON l.id_licencia = da.id_licencia
        LEFT JOIN copias_seguridad cs ON l.id_licencia = cs.id_licencia
        ORDER BY s.area_departamento, l.nombre_licencia
    ''')
    rows = cursor.fetchall()
    cerrar_conexion(cursor, conexion)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Licencias"

    # Definir estilos
    header_fill = PatternFill(start_color="09524D", end_color="09524D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Encabezados
    headers = [
        'ID Licencia', 'Nombre Licencia', 'Descripción', 'Tipo', 'Licenciamiento', 'Proveedor',
        'Frecuencia Uso', 'Fecha Vencimiento', 'ID Solicitud', 'Área', 'Encargado', 'Solicitante',
        'Fecha Solicitud', 'Cantidad Usuarios', 'Maneja Roles', 'Gestor Roles', 'Gestor Usuarios',
        'Maneja Info Sensible', 'Tipo Info Sensible', 'Validación Compliance', 'Realiza Copias',
        'Tipo Copia', 'Frecuencia Copia', 'Ubicación Copia', 'Tiene Integraciones', 'Detalle Integraciones',
        'Impacto', 'Procesos Relacionados', 'Contacto Proveedor'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Agregar datos
    for row_num, row in enumerate(rows, 2):
        # Valores por defecto y lógica condicional
        fecha_venc = row.get('fecha_vencimiento') if row.get('fecha_vencimiento') else 'NO'

        realiza_copias = (row.get('realiza_copias') or '').strip()
        tipo_copia = row.get('tipo_copia') or ''
        frecuencia_copia = row.get('frecuencia_copia') or ''
        ubicacion_copia = row.get('ubicacion_copia') or ''
        if realiza_copias.lower() != 'si':
            tipo_copia = 'No aplica'
            frecuencia_copia = 'No aplica'
            ubicacion_copia = 'No aplica'

        maneja_info = (row.get('maneja_info_sensible') or '').strip()
        tipo_info = row.get('tipo_info_sensible') or ''
        if maneja_info.lower() != 'si':
            tipo_info = 'No aplica'

        maneja_roles = (row.get('maneja_roles') or '').strip()
        gestor_roles = row.get('gestor_roles') or ''
        gestor_usuarios = row.get('gestor_usuarios') or ''
        if maneja_roles.lower() != 'si':
            gestor_roles = 'No aplica'
            gestor_usuarios = 'No aplica'

        tiene_integraciones = (row.get('tiene_integraciones') or '').strip()
        detalle_integraciones = row.get('detalle_integraciones') or ''
        if tiene_integraciones.lower() != 'si':
            detalle_integraciones = 'No aplica'

        # Asignar valores a celdas
        # Usar numeración consecutiva en el Excel en lugar del id real de la base de datos
        ws.cell(row=row_num, column=1).value = row_num - 1
        ws.cell(row=row_num, column=2).value = row.get('nombre_licencia') or ''
        ws.cell(row=row_num, column=3).value = row.get('descripcion') or ''
        ws.cell(row=row_num, column=4).value = row.get('tipo_herramienta') or ''
        ws.cell(row=row_num, column=5).value = row.get('licenciamiento') or ''
        ws.cell(row=row_num, column=6).value = row.get('proveedor') or ''
        ws.cell(row=row_num, column=7).value = row.get('frecuencia_uso') or ''
        ws.cell(row=row_num, column=8).value = fecha_venc
        ws.cell(row=row_num, column=9).value = row.get('id_solicitud')
        ws.cell(row=row_num, column=10).value = row.get('area_departamento') or ''
        ws.cell(row=row_num, column=11).value = row.get('jefe_area') or ''
        ws.cell(row=row_num, column=12).value = row.get('solicitante') or ''
        ws.cell(row=row_num, column=13).value = row.get('fecha_solicitud') or ''
        ws.cell(row=row_num, column=14).value = row.get('cantidad_usuarios') or ''
        ws.cell(row=row_num, column=15).value = maneja_roles or ''
        ws.cell(row=row_num, column=16).value = gestor_roles
        ws.cell(row=row_num, column=17).value = gestor_usuarios
        ws.cell(row=row_num, column=18).value = maneja_info or ''
        ws.cell(row=row_num, column=19).value = tipo_info
        ws.cell(row=row_num, column=20).value = row.get('validacion_compliance') or ''
        ws.cell(row=row_num, column=21).value = realiza_copias or ''
        ws.cell(row=row_num, column=22).value = tipo_copia
        ws.cell(row=row_num, column=23).value = frecuencia_copia
        ws.cell(row=row_num, column=24).value = ubicacion_copia
        ws.cell(row=row_num, column=25).value = tiene_integraciones or ''
        ws.cell(row=row_num, column=26).value = detalle_integraciones
        ws.cell(row=row_num, column=27).value = row.get('impacto_falla') or ''
        ws.cell(row=row_num, column=28).value = row.get('procesos_relacionados') or ''
        ws.cell(row=row_num, column=29).value = row.get('contacto_proveedor') or ''

        # Aplicar estilos a cada fila
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = left_alignment

    # Ajustar ancho de columnas
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 18

    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'licencias_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/licencia/<int:id>/eliminar', methods=['POST'])
def eliminar_licencia(id):
    next_view = request.form.get('next', 'global')
    

    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)
    cursor.execute('SELECT id_solicitud FROM licencias WHERE id_licencia = %s', (id,))
    row = cursor.fetchone()
    solicitud_id = row['id_solicitud'] if row else None

    cursor.execute('DELETE FROM datos_adicionales WHERE id_licencia = %s', (id,))
    cursor.execute('DELETE FROM copias_seguridad WHERE id_licencia = %s', (id,))
    cursor.execute('DELETE FROM licencias WHERE id_licencia = %s', (id,))
    conexion.commit()
    cerrar_conexion(cursor, conexion)

    if next_view == 'resumen' and solicitud_id:
        return redirect(f'/solicitud/{solicitud_id}/resumen')
    if next_view == 'vista':
        return redirect('/vista')
    if next_view == 'solicitudes':
        return redirect('/solicitudes')
    return redirect('/vista_global')

# ---------------------------

@app.route('/detalle/<int:id>')
def detalle(id):
    conexion = obtener_conexion()
    cursor = conexion.cursor(dictionary=True)

    cursor.execute('''
    SELECT 
        s.id_solicitud,
        s.area_departamento,
        s.jefe_area,
        s.solicitante,
        s.fecha_solicitud,

        l.id_licencia,
        l.nombre_licencia,
        l.descripcion,
        l.tipo_herramienta,
        l.licenciamiento,
        l.proveedor,
        l.frecuencia_uso,
        l.fecha_vencimiento,

        d.cantidad_usuarios,
        d.maneja_roles,
        d.gestor_roles,
        d.gestor_usuarios,
        d.maneja_info_sensible,
        d.tipo_info_sensible,
        d.validacion_compliance,

        c.realiza_copias,
        c.tipo_copia,
        c.frecuencia_copia,
        c.ubicacion_copia,
        c.tiene_integraciones,
        c.detalle_integraciones,
        c.impacto_falla,
        c.procesos_relacionados,
        c.contacto_proveedor

    FROM licencias l
    JOIN solicitudes s ON l.id_solicitud = s.id_solicitud
    LEFT JOIN datos_adicionales d ON l.id_licencia = d.id_licencia
    LEFT JOIN copias_seguridad c ON l.id_licencia = c.id_licencia

    WHERE l.id_licencia = %s
    ''', (id,))

    detalle = cursor.fetchone()
    cerrar_conexion(cursor, conexion)

    next_view = request.args.get('next')
    if next_view == 'resumen' and detalle:
        return_url = f"/solicitud/{detalle['id_solicitud']}/resumen"
    elif next_view == 'global':
        return_url = '/vista_global'
    elif next_view == 'vista':
        return_url = '/vista'
    else:
        return_url = '/vista'

    return render_template('detalle.html', d=detalle, return_url=return_url, next_view=next_view)

# ---------------------------
# ACTUALIZAR LICENCIA
# ---------------------------
@app.route('/actualizar_licencia/<int:id>', methods=['POST'])
def actualizar_licencia(id):
    conexion = obtener_conexion()
    cursor = conexion.cursor()
    
    # Obtener los datos del formulario
    data = request.form.to_dict()
    
    # Mapeo de campos a tablas
    licencias_fields = ['nombre_licencia', 'descripcion', 'tipo_herramienta', 'licenciamiento', 'proveedor', 'frecuencia_uso', 'fecha_vencimiento']
    datos_adicionales_fields = ['cantidad_usuarios', 'maneja_roles', 'gestor_roles', 'gestor_usuarios', 'maneja_info_sensible', 'tipo_info_sensible', 'validacion_compliance']
    copias_seguridad_fields = ['realiza_copias', 'tipo_copia', 'frecuencia_copia', 'ubicacion_copia', 'tiene_integraciones', 'detalle_integraciones', 'impacto_falla', 'procesos_relacionados', 'contacto_proveedor']
    
    # Actualizar tabla licencias
    update_fields_lic = {k: v for k, v in data.items() if k in licencias_fields}
    if update_fields_lic:
        set_clause = ', '.join([f'{k}=%s' for k in update_fields_lic.keys()])
        values = list(update_fields_lic.values()) + [id]
        sql_lic = f'UPDATE licencias SET {set_clause} WHERE id_licencia=%s'
        cursor.execute(sql_lic, values)
    
    # Actualizar tabla datos_adicionales
    update_fields_dat = {k: v for k, v in data.items() if k in datos_adicionales_fields}
    if update_fields_dat:
        set_clause = ', '.join([f'{k}=%s' for k in update_fields_dat.keys()])
        values = list(update_fields_dat.values()) + [id]
        sql_dat = f'UPDATE datos_adicionales SET {set_clause} WHERE id_licencia=%s'
        cursor.execute(sql_dat, values)
    
    # Actualizar tabla copias_seguridad
    update_fields_cop = {k: v for k, v in data.items() if k in copias_seguridad_fields}
    if update_fields_cop:
        set_clause = ', '.join([f'{k}=%s' for k in update_fields_cop.keys()])
        values = list(update_fields_cop.values()) + [id]
        sql_cop = f'UPDATE copias_seguridad SET {set_clause} WHERE id_licencia=%s'
        cursor.execute(sql_cop, values)
    
    conexion.commit()
    cerrar_conexion(cursor, conexion)
    
    # Devolver JSON para fetch
    return jsonify({'success': True, 'message': 'Actualizado correctamente'})

# ---------------------------
# EJECUCIÓN
# ---------------------------

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
